from io import BytesIO
from datetime import timedelta, datetime
from openpyxl import Workbook
from website.database.models import Users, Trackers
from .data import ListTimes
from datetime import timedelta, datetime


def Excel(users, month, year):
    myio = BytesIO()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Danh sách nhân viên"
    ws1['B1'] = "Danh sách nhân viên"
    ws1['B2'] = "Chấm công tháng " + month
    ws1['A3'] = 'Họ và tên'
    ws1['B3'] = 'Ngày sinh'
    ws1['C3'] = 'Chức vụ'
    i = 4
    total = timedelta(hours=0, minutes=0)
    for user in users:
        employee = Users.objects.get(pk=user)
        ws1['A' + str(i)] = employee['lastName'] + ' ' + employee['firstName']
        # ws1['B' + str(i)] = employee['birth']
        ws1['C' + str(i)] = employee['affiliation']
        ws = wb.create_sheet(employee['displayName'])
        times = Trackers.objects(user=user)
        ws['A1'] = "Tên nhân viên: "
        ws['A2'] = "Chức vụ: "
        ws['C2'] = "Ngày sinh: "
        ws['A3'] = "Ngày"
        ws['B3'] = "Check in"
        ws['C3'] = "Check out"
        ws['D3'] = 'Tổng thời gian'
        ws['B1'] = employee.firstName + ' ' + employee.lastName
        ws['B2'] = employee.affiliation
        # ws['D2'] = employee.birth
        list_time = ListTimes(times, month, year)
        if(list_time != None):
            j = 5
            for time in list_time:
                ws['A' + str(j)] = time["id"]
                ws['B' + str(j)] = time["checkIn"]
                ws['C' + str(j)] = time["checkOut"]
                ws['D' + str(j)] = time["total"]
                a = time["total"].split(':')  # Cắt time để lấy ngày và giờ
                t = timedelta(hours=int(a[0]), minutes=int(a[1]))
                total += t
                j += 1
            ws['A' + str(j)] = "Tổng thời gian làm"
            ws['D' + str(j)] = total
            i += 1
            total = timedelta(hours=0, minutes=0)
    wb.save(myio)
    myio.seek(0)
    return myio
